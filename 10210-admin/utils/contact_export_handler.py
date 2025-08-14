"""
Contact Export Handler - Handles exporting WhatsApp contacts to JSON, Excel, and CSV formats
Uses the same 16-column format as group exports for consistency
"""

import json
import logging
import os
import csv
from datetime import datetime
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

logger = logging.getLogger(__name__)

class ContactExportHandler:
    """Handles exporting WhatsApp contacts to various formats"""
    
    def __init__(self, export_dir: str = "static/exports"):
        self.export_dir = export_dir
        os.makedirs(export_dir, exist_ok=True)
    
    def export_contacts(
        self, 
        contacts: List[Dict[str, Any]], 
        session_name: str
    ) -> Dict[str, Any]:
        """
        Export WhatsApp contacts to JSON, Excel and CSV formats
        Uses same 16-column structure as group exports for consistency
        
        Args:
            contacts: List of contact details from WAHA API
            session_name: WhatsApp session name
            
        Returns:
            Dictionary with export results and file paths
        """
        try:
            # Filter contacts: only include saved contacts OR contacts with chat history
            # Skip groups and @lid contacts
            filtered_contacts = []
            for contact in contacts:
                # Skip groups
                if contact.get('isGroup', False) or '@g.us' in str(contact.get('id', '')):
                    continue
                
                # Skip @lid contacts (these are not real phone numbers)
                if '@lid' in str(contact.get('id', '')):
                    continue
                
                # Include if it's a saved contact OR has chat history
                # WAHA provides lastMessage or type='in' for contacts with chats
                has_chat = contact.get('type') == 'in' or contact.get('lastMessage') is not None
                is_my_contact = contact.get('isMyContact', False)
                
                if is_my_contact or has_chat:
                    filtered_contacts.append(contact)
            
            # Generate filename with timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            base_filename = f"contacts_{session_name}_{timestamp}"
            
            # Prepare data for export using 16-column format
            export_data = self._prepare_export_data(filtered_contacts, session_name)
            
            # Export to JSON
            json_path = self._export_to_json(export_data, base_filename)
            
            # Export to Excel
            excel_path = self._export_to_excel(export_data, base_filename)
            
            # Export to CSV
            csv_path = self._export_to_csv(export_data, base_filename)
            
            # Generate download URLs (relative to static directory)
            json_url = f"/exports/{os.path.basename(json_path)}"
            excel_url = f"/exports/{os.path.basename(excel_path)}"
            csv_url = f"/exports/{os.path.basename(csv_path)}"
            
            return {
                "success": True,
                "json_url": json_url,
                "excel_url": excel_url,
                "csv_url": csv_url,
                "contact_count": len(contacts),
                "session_name": session_name,
                "exported_at": datetime.now().isoformat()
            }
            
        except Exception as e:
            logger.error(f"Failed to export contacts: {str(e)}")
            raise
    
    def _prepare_export_data(self, contacts: List[Dict], session_name: str) -> Dict[str, Any]:
        """
        Prepare contact data for export in 16-column format
        Matches the structure used in group exports
        """
        processed_contacts = []
        
        for contact in contacts:
            # Use the 'number' field from WAHA which is clean
            phone_number = contact.get('number', '')
            
            # Skip if no valid phone number
            if not phone_number:
                continue
            
            # Extract country code first
            country_code = ''
            country_name = 'Unknown'
            if phone_number:
                # Simple country code extraction (first 1-3 digits)
                if phone_number.startswith('1') and len(phone_number) == 11:
                    country_code = '+1'
                    country_name = 'United States/Canada'
                elif phone_number.startswith('234') and len(phone_number) >= 13:
                    country_code = '+234'
                    country_name = 'Nigeria'
                elif phone_number.startswith('91') and len(phone_number) >= 12:
                    country_code = '+91'
                    country_name = 'India'
                elif phone_number.startswith('44'):
                    country_code = '+44'
                    country_name = 'United Kingdom'
                # Add more country codes as needed
            
            # Format phone with space after country code (same as group export)
            formatted_phone = phone_number
            if phone_number:
                if phone_number.startswith('1') and len(phone_number) == 11:
                    formatted_phone = f"+1 {phone_number[1:]}"
                elif phone_number.startswith('234') and len(phone_number) >= 13:
                    formatted_phone = f"+234 {phone_number[3:]}"
                elif phone_number.startswith('91') and len(phone_number) >= 12:
                    formatted_phone = f"+91 {phone_number[2:]}"
                elif phone_number.startswith('44'):
                    formatted_phone = f"+44 {phone_number[2:]}"
                elif not phone_number.startswith('+'):
                    # Default: just add + if missing
                    formatted_phone = f"+{phone_number}"
            
            # Get last message info if available
            last_msg = contact.get('lastMessage', {})
            last_msg_text = ''
            last_msg_date = ''
            last_msg_type = ''
            last_msg_status = ''
            
            if last_msg:
                last_msg_text = last_msg.get('body', '') or '[Media]'
                if last_msg.get('timestamp'):
                    try:
                        last_msg_date = datetime.fromtimestamp(last_msg['timestamp']).strftime('%Y-%m-%d %H:%M:%S')
                    except:
                        last_msg_date = ''
                last_msg_type = last_msg.get('type', '')
                last_msg_status = 'sent' if last_msg.get('fromMe') else 'received'
            
            # Build 16-column structure matching group export
            processed_contact = {
                'phone_number': formatted_phone,  # Use formatted version with space
                'formatted_phone': formatted_phone,
                'country_code': country_code,
                'country_name': country_name,
                'saved_name': contact.get('name', ''),
                'public_name': contact.get('pushname', ''),
                'is_my_contact': contact.get('isMyContact', False),
                'is_business': contact.get('isBusiness', False),
                'is_blocked': contact.get('isBlocked', False),
                'is_admin': False,  # Not applicable for contacts
                'is_super_admin': False,  # Not applicable for contacts
                'labels': ', '.join(contact.get('labels', [])) if contact.get('labels') else '',
                'last_msg_text': last_msg_text[:100] if last_msg_text else '',
                'last_msg_date': last_msg_date,
                'last_msg_type': last_msg_type,
                'last_msg_status': last_msg_status
            }
            
            processed_contacts.append(processed_contact)
        
        return {
            "session_name": session_name,
            "export_date": datetime.now().isoformat(),
            "total_contacts": len(contacts),
            "contacts": processed_contacts
        }
    
    def _export_to_json(self, data: Dict[str, Any], filename: str) -> str:
        """Export data to JSON format"""
        json_path = os.path.join(self.export_dir, f"{filename}.json")
        
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        
        logger.info(f"Exported to JSON: {json_path}")
        return json_path
    
    def _export_to_excel(self, data: Dict[str, Any], filename: str) -> str:
        """Export data to Excel format with formatting"""
        excel_path = os.path.join(self.export_dir, f"{filename}.xlsx")
        
        # Create workbook and sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Contacts"
        
        # Add column headers (16 columns matching group export)
        headers = [
            'phone_number', 'formatted_phone', 'country_code', 'country_name',
            'saved_name', 'public_name', 'is_my_contact', 'is_business', 
            'is_blocked', 'is_admin', 'is_super_admin', 'labels',
            'last_msg_text', 'last_msg_date', 'last_msg_type', 'last_msg_status'
        ]
        
        # Write headers with formatting
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Write contact data starting from row 2
        for row_idx, contact in enumerate(data['contacts'], 2):
            ws.cell(row=row_idx, column=1, value=contact['phone_number'])
            ws.cell(row=row_idx, column=2, value=contact['formatted_phone'])
            ws.cell(row=row_idx, column=3, value=contact['country_code'])
            ws.cell(row=row_idx, column=4, value=contact['country_name'])
            ws.cell(row=row_idx, column=5, value=contact['saved_name'])
            ws.cell(row=row_idx, column=6, value=contact['public_name'])
            ws.cell(row=row_idx, column=7, value='Yes' if contact['is_my_contact'] else 'No')
            ws.cell(row=row_idx, column=8, value='Yes' if contact['is_business'] else 'No')
            ws.cell(row=row_idx, column=9, value='Yes' if contact['is_blocked'] else 'No')
            ws.cell(row=row_idx, column=10, value='Yes' if contact['is_admin'] else 'No')
            ws.cell(row=row_idx, column=11, value='Yes' if contact['is_super_admin'] else 'No')
            ws.cell(row=row_idx, column=12, value=contact['labels'])
            ws.cell(row=row_idx, column=13, value=contact['last_msg_text'])
            ws.cell(row=row_idx, column=14, value=contact['last_msg_date'])
            ws.cell(row=row_idx, column=15, value=contact['last_msg_type'])
            ws.cell(row=row_idx, column=16, value=contact['last_msg_status'])
            
            # Alternate row coloring
            if row_idx % 2 == 0:
                for col in range(1, 17):
                    ws.cell(row=row_idx, column=col).fill = PatternFill(
                        start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
                    )
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # Save workbook
        wb.save(excel_path)
        logger.info(f"Exported to Excel: {excel_path}")
        return excel_path
    
    def _export_to_csv(self, data: Dict[str, Any], filename: str) -> str:
        """Export data to CSV format - compatible with campaign processor"""
        csv_path = os.path.join(self.export_dir, f"{filename}.csv")
        
        # Define headers using same format as group export
        headers = [
            'phone_number', 'formatted_phone', 'country_code', 'country_name',
            'saved_name', 'public_name', 'is_my_contact', 'is_business', 
            'is_blocked', 'is_admin', 'is_super_admin', 'labels',
            'last_msg_text', 'last_msg_date', 'last_msg_type', 'last_msg_status'
        ]
        
        with open(csv_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=headers)
            writer.writeheader()
            
            # Write contact data
            for contact in data['contacts']:
                # Ensure phone_number has + prefix for campaign compatibility
                row = {
                    'phone_number': contact.get('formatted_phone', ''),  # Use formatted_phone with +
                    'formatted_phone': contact.get('formatted_phone', ''),
                    'country_code': f"+{contact.get('country_code', '')}" if contact.get('country_code', '') else '',
                    'country_name': contact.get('country_name', ''),
                    'saved_name': contact.get('saved_name', ''),
                    'public_name': contact.get('public_name', ''),
                    'is_my_contact': 'true' if contact.get('is_my_contact') else 'false',
                    'is_business': 'true' if contact.get('is_business') else 'false',
                    'is_blocked': 'true' if contact.get('is_blocked') else 'false',
                    'is_admin': 'true' if contact.get('is_admin') else 'false',
                    'is_super_admin': 'true' if contact.get('is_super_admin') else 'false',
                    'labels': contact.get('labels', ''),
                    'last_msg_text': contact.get('last_msg_text', ''),
                    'last_msg_date': contact.get('last_msg_date', ''),
                    'last_msg_type': contact.get('last_msg_type', ''),
                    'last_msg_status': contact.get('last_msg_status', '')
                }
                writer.writerow(row)
        
        logger.info(f"Exported to CSV: {csv_path}")
        return csv_path